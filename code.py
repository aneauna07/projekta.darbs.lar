import PyPDF2
import os
from openpyxl import Workbook, load_workbook
from datetime import timedelta
import sys

input = input("Klienta Vārds Uzvārds:")    #klients ievada vārdu un uzvārdu

klienti=[]

with open("klienti.csv", encoding="utf-8") as f:
    next(f)
    for line in f:                                              #informācijas iegūšana no csv faila
        row=line.rstrip().split(";")
        klienti.append(row)


file="inventars.pdf"
isExisting = os.path.exists(file)

if isExisting == False:
     print("0")                        
     exit()
else:
     if file!="":
          inventars=[]
          cena=[]
          pdf_file=PyPDF2.PdfReader(open(file,"rb"))
          number_of_pages=len(pdf_file.pages)
          page1=pdf_file.pages[0]                        #no pdf faila tiek izgūts teksts
          text1=page1.extract_text()
          pos1=text1.find("min:")
          pos2=text1.find("noma(1)")
          pos3=text1.find("(1):")
          pos4=text1.find("Slēpes")
          pos5=text1.find("9.00")
          pos6=text1.find("noma(2)")
          pos7=text1.find("(2):")
          pos8=text1.find("Sniega")
          pos9=text1.find("noma(2):  5.00")
          pos10=text1.find("noma(3):")
          pos11=text1.find("(3):")
          pos12=text1.find("Zābaki")
          pos13=text1.find("noma(3):  5.00")
          pos14=text1.find("noma(4)")
          pos15=text1.find("(4):")
          pos16=text1.find("Nūjas")
          pos17=text1.find("4.00")
          pos18=text1.find("noma(5)")
          pos19=text1.find("(5):")
          pos20=text1.find("Ķivere")
          pos21=text1.find("2.00")
          pos22=text1.find("noma(6)")
          pos23=text1.find("(6):")
          komplekts = text1[pos1+5:pos2].rstrip()
          cena1 = text1[pos2+12:pos4-2]
          slepes = text1[pos5+6:pos6].rstrip()
          cena2 = text1[pos7+8:pos8-1]
          delis = text1[pos9+140:pos10].rstrip() 
          cena3 = text1[pos11+8:pos12-1]
          zabaki = text1[pos13+196:pos14].rstrip() 
          cena4 = text1[pos15+8:pos16-1] 
          nujas = text1[pos17+5:pos18].rstrip() 
          cena5 = text1[pos19+8:pos20-1]
          kivere = text1[pos21+5:pos22].rstrip()
          cena6 = text1[pos23+8:pos23+12] 
          inventars.append(komplekts)
          cena.append(cena1)
          inventars.append(slepes)
          cena.append(cena2)
          inventars.append(delis)
          cena.append(cena3)
          inventars.append(zabaki)
          cena.append(cena4)
          inventars.append(nujas)
          cena.append(cena5)
          inventars.append(kivere)
          cena.append(cena6)

inventars1 = []

inventars1 = [[inventars[i], cena[i]] for i in range(min(len(inventars), len(cena)))]     #informācija no pdf faila apvienota vienā


wb=load_workbook('noma1.xlsx')
noma=[]
ws=wb.active
max_row=ws.max_row                      
for row in range(2,200):                 
    r_data=[]
    a_value=ws['A'+str(row)].value            
    b_value=ws['B'+str(row)].value 
    c_value=ws['C'+str(row)].value     
    r_data.append(a_value)
    r_data.append(b_value)                         #informācijas iegūšana no xlsx faila
    r_data.append(c_value)       
    noma.append(r_data)


code = 0

for x in range(len(klienti)):
    i = klienti[x][1]
    if i == input:
        code = klienti[x][0]
        break 
                              #ievaditais vards un uzvards tiek meklets klientu datubazes sarakstaa

laiks= []
inv = []

for x in range(len(noma)):
    row=[]
    i = noma[x][0]
    if i == code:
        laiks.append(noma[x][1])
        inv.append(noma[x][2])
    else:
        print("0.00€")
        print("00:00:00")
        sys.exit()
     
apvienojums = {item[0]: item[1] for item in inventars1}

rezultats = [apvienojums[name] for name in inv if name in apvienojums]          

laiks_maksa = [[laiks[i], rezultats[i]] for i in range(min(len(laiks), len(rezultats)))]    #izveidots kopigs saraksts vajadzigajai informacijai


indekss = 1
float_inv1 = [[float(element) if index == indekss else element for index, element in enumerate(row)] for row in inventars1]   #cenas parveidotas par float

cenamin =[]
for row in float_inv1:
    cenaa = row[indekss] / 60.00
    cenamin.append(cenaa)                  #aprekinata cena par vienu min

laiks_maksa_min = [[cenamin[i], laiks_maksa[i]] for i in range(min(len(cenamin), len(laiks_maksa)))]

sum = 0
total_hours = 0
total_minutes = 0
stundas_kop = 0
min_kop = 0
sec_kop = 0

for item in laiks_maksa_min:
    stundas = item[1][0].hour
    minutes = item[1][0].minute
    sec = item[1][0].second         #izgutas laika vienibas

    total_hours += stundas * float(item[1][1])
    total_minutes += minutes * item[0]
    sum = total_hours + total_minutes                     #aprekinata kopejaa maksajuma summa

    stundas_kop = stundas_kop + stundas
    min_kop = min_kop + minutes
    sec_kop = sec_kop + sec              #aprekinats kopejais laiks

    koplaiks = timedelta(hours=stundas_kop, minutes=min_kop, seconds=sec_kop)   #vizuali attelots kopejais laiks

formats_sum = "{:.2f}".format(sum)         #ar diviem cipariem aiz komata

if code != 0:
    print(formats_sum+ "€")
    print(str(koplaiks))
else:
    print("0.00€")
    print("00:00:00")
